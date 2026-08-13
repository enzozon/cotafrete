"""Testes da camada pura: verificam se o dado certo vai para o campo certo.
Nenhum destes testes toca a internet ou o site da Della Volpe."""

from __future__ import annotations

from decimal import Decimal

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from carriers.base import Severidade
from carriers.dellavolpe import mapping as dv
from carriers.dellavolpe.planilha import gerar_planilha_volumes
from core.models import (
    CotacaoRequest, Local, Mercadoria, NotaFiscal, Parte, Servico,
    Solicitante, StatusCotacao, Volume, cnpj_valido,
)

CNPJ_A = "11.222.333/0001-81"
CNPJ_B = "45.723.174/0001-10"
CNPJ_C = "61.139.432/0001-72"  # o da própria Della Volpe, público no site


def montar(**over) -> CotacaoRequest:
    base = dict(
        solicitante=Solicitante(nome="Teste da Silva", email="t@ex.com",
                                whatsapp="27999887766"),
        servico=Servico.FRACIONADO_LTL,
        origem=Local(uf="es", cidade="Vitória"),
        destino=Local(uf="SP", cidade="São Paulo"),
        remetente=Parte(cnpj=CNPJ_A),
        destinatario=Parte(cnpj=CNPJ_B),
        pagador_frete=Parte(cnpj=CNPJ_C),
        volumes=[Volume(qtd=2, comprimento_cm=Decimal(100), largura_cm=Decimal(50),
                        altura_cm=Decimal(40), peso_kg=Decimal(10))],
        mercadoria=Mercadoria(tipo_material="Peças metálicas"),
        nota_fiscal=NotaFiscal(valor_total=Decimal(25_000)),
    )
    base.update(over)
    return CotacaoRequest(**base)


# ------------------------------------------------------------- fundamentos
def test_cnpj_mod11():
    assert cnpj_valido(CNPJ_A)
    assert cnpj_valido(CNPJ_C)
    assert not cnpj_valido("11.222.333/0001-80")
    assert not cnpj_valido("11111111111111")


def test_cnpj_invalido_rejeitado_no_modelo():
    with pytest.raises(ValidationError):
        montar(remetente=Parte(cnpj="00.000.000/0000-00"))


def test_uf_normalizada_para_maiuscula():
    assert montar().origem.uf == "ES"


# --------------------------------------------------------------- derivados
def test_peso_e_quantidade_multiplicam_pela_qtd():
    r = montar()
    assert r.peso_total_kg == Decimal(20)      # 10 kg x 2
    assert r.quantidade_volumes == 2


def test_cubagem_em_m3_nao_cm3():
    r = montar()
    # 100 x 50 x 40 = 200.000 cm³ x 2 = 400.000 cm³ = 0,4 m³
    assert r.cubagem_m3 == Decimal("0.4")


def test_cubagem_soma_volumes_heterogeneos():
    r = montar(volumes=[
        Volume(qtd=1, comprimento_cm=Decimal(100), largura_cm=Decimal(50),
               altura_cm=Decimal(40), peso_kg=Decimal(10)),
        Volume(qtd=1, comprimento_cm=Decimal(80), largura_cm=Decimal(40),
               altura_cm=Decimal(30), peso_kg=Decimal(8)),
    ])
    assert r.peso_total_kg == Decimal(18)
    assert r.quantidade_volumes == 2
    assert r.cubagem_m3 == Decimal("0.296")   # 0,2 + 0,096
    assert r.tem_medidas_distintas is True


@pytest.mark.parametrize("comp, larg, alt, cubado_na_proposta", [
    (40, 30, 20, Decimal("7.20")),      # proposta 13320/26
    (50, 40, 30, Decimal("18.00")),     # proposta 13322/26
    (60, 40, 40, Decimal("28.80")),     # proposta 13324/26
    (80, 60, 50, Decimal("72.00")),     # proposta 13326/26
])
def test_fator_300_confere_com_as_propostas_reais(comp, larg, alt,
                                                  cubado_na_proposta):
    """Fator CONFIRMADO em 13/08/2026, não mais presumido.

    Quatro propostas da Della Volpe declararam o peso cubado que elas mesmas
    calcularam, para cargas que nós enviamos. Dividindo pelo volume em m³ dá
    300 nas quatro. Era a última suposição grande que restava no projeto.
    """
    req = montar(volumes=[Volume(qtd=1, comprimento_cm=Decimal(comp),
                                 largura_cm=Decimal(larg),
                                 altura_cm=Decimal(alt), peso_kg=Decimal(1))])

    assert req.peso_cubado_kg(dv.FATOR_CUBAGEM) == cubado_na_proposta


def test_peso_cubado_usa_fator_da_transportadora():
    r = montar()
    assert r.peso_cubado_kg(dv.FATOR_CUBAGEM) == Decimal("120.0")  # 0,4 x 300


# -------------------------------------------------------------- formatação
@pytest.mark.parametrize("entrada,esperado", [
    (Decimal(25_000), "25.000,00"),
    (Decimal("1234.5"), "1.234,50"),
    (Decimal("999.99"), "999,99"),
    (Decimal(1_500_000), "1.500.000,00"),
])
def test_valor_em_formato_brasileiro(entrada, esperado):
    assert dv.num_br(entrada) == esperado


def test_peso_sem_separador_de_milhar():
    assert dv.peso_br(Decimal(1500)) == "1500"
    assert dv.peso_br(Decimal("1500.5")) == "1500,5"


def test_a_palavra_sucesso_sozinha_nao_prova_envio():
    """Medido em 13/08/2026: a página da Della Volpe já contém "sucesso" no
    HTML ANTES de qualquer submissão.

    O detector antigo procurava obrigado/sucesso/enviad/recebemos no HTML
    inteiro, então dava "aguardando_retorno" para qualquer página — inclusive
    uma em que o envio falhou. Cinco cotações voltaram com esse status sem
    que ninguém pudesse afirmar que saíram."""
    html_sem_envio = """<html><body>
        <div class="case-sucesso">Casos de sucesso</div>
        <div class="wpcf7-response-output" aria-hidden="true"></div>
    </body></html>"""

    res = dv.normalizar_resposta(html_sem_envio)
    assert res.status is StatusCotacao.ERRO
    assert "não identificada" in res.erro


def test_confirmacao_real_do_site_e_reconhecida():
    """HTML exato do site, mandado pelo Enzo em 13/08/2026.

    Repare no que NÃO tem: a classe wpcf7-mail-sent-ok. Este site mantém a div
    como wpcf7-response-output aria-hidden="true" e só troca o TEXTO de dentro.
    A correção anterior, que exigia a classe, teria recusado toda confirmação
    verdadeira — trocar um falso positivo por um falso negativo."""
    html = ('<div class="wpcf7-response-output" aria-hidden="true">Olá Enzo '
            'Zon. Agradecemos a sua mensagem. Em breve retornaremos seu '
            'contato.</div>')

    res = dv.normalizar_resposta(html)
    assert res.status is StatusCotacao.AGUARDANDO_RETORNO
    assert res.valor_frete is None      # o preço só chega por e-mail
    assert res.erro is None


def test_bloqueio_antispam_tem_status_proprio():
    """Resposta real do site em 13/08/2026, com o envio já funcionando.

    O CF7 barrou a submissão como spam — nenhum e-mail foi gerado. Cair no
    ERRO genérico esconderia a causa: o operador ficaria procurando bug no
    preenchimento enquanto o problema é reputação do remetente."""
    html = ('<div class="wpcf7-response-output" aria-hidden="true">'
            'A submissão mencionou-se como spam. Clique em "Pedir orçamento" '
            'novamente</div>')

    res = dv.normalizar_resposta(html)
    assert res.status is StatusCotacao.INTERVENCAO_NECESSARIA
    assert "spam" in res.motivo_recusa.lower()
    assert res.valor_frete is None


def test_div_de_resposta_vazia_nao_e_envio():
    """Antes do submit a div existe e está VAZIA. É esse o estado que voltou
    nas cinco cotações que não geraram e-mail nenhum."""
    html = '<div class="wpcf7-response-output" aria-hidden="true"></div>'
    assert dv.normalizar_resposta(html).status is StatusCotacao.ERRO


def test_secao_casos_de_sucesso_da_pagina_nao_conta():
    """O texto que interessa é o de DENTRO da div de resposta, não o da
    página. Foi a seção "Casos de sucesso" que enganou o detector antigo."""
    html = ('<div class="case-sucesso">Casos de sucesso</div>'
            '<div class="wpcf7-response-output" aria-hidden="true"></div>')
    assert dv.normalizar_resposta(html).status is StatusCotacao.ERRO


def test_erro_declarado_pelo_contact_form_7_nao_vira_sucesso():
    html_falhou = ('<div class="wpcf7-response-output wpcf7-mail-sent-ng">'
                   'Ocorreu um erro ao tentar enviar sua mensagem.</div>')
    assert dv.normalizar_resposta(html_falhou).status is StatusCotacao.ERRO


def test_medidas_usam_uma_casa_decimal_por_causa_da_mascara():
    """Medido no site em produção: o campo de medida tem máscara que reserva
    UMA casa decimal. Digitar '100' vira '10,0' — a carga seria cotada 10x
    menor, em silêncio. O próprio placeholder declara o formato esperado:
    'Comprimento (ex. 12,5m = 1.250,0cm)'.

    peso_br() é formatador de PESO e não serve aqui: ele corta a casa decimal."""
    assert dv.medida_br(Decimal(100)) == "100,0"
    assert dv.medida_br(Decimal(40)) == "40,0"
    assert dv.medida_br(Decimal(1250)) == "1.250,0"
    assert dv.medida_br(Decimal("12.5")) == "12,5"


# ----------------------------------------------------------------- payload
def test_payload_mapeia_todos_os_campos_obrigatorios():
    campos = dv.campos_do_formulario(dv.preparar_payload(montar()))
    for spec in dv.campos_obrigatorios(montar()):
        assert spec.nome in campos, f"campo obrigatório ausente: {spec.nome}"


def test_payload_valores_exatos():
    p = dv.preparar_payload(montar())
    assert p["Nome completo"] == "Teste da Silva"
    assert p["WhatsApp"] == "(27) 99988-7766"
    assert p["Qual o serviço que você procura?"] == "Fracionado -LTL"
    assert p["CNPJ - Remetente"] == "11.222.333/0001-81"
    assert p["CNPJ - Destinatário"] == "45.723.174/0001-10"
    assert p["CNPJ da empresa que pagará o frete"] == "61.139.432/0001-72"
    assert p["Selecione o estado de origem"] == "ES"
    assert p["Peso total"] == "20"
    assert p["Quantidade de Volumes"] == "2"
    assert p["Comprimento"] == "100,0"    # máscara de 1 casa; ver medida_br()
    assert p["Largura"] == "50,0"
    assert p["Altura"] == "40,0"
    assert p["Valor total da nota fiscal"] == "25.000,00"


def test_cnpjs_nao_se_misturam():
    """Regressão: os três CNPJs são campos diferentes e não podem trocar."""
    p = dv.preparar_payload(montar())
    tres = {p["CNPJ - Remetente"], p["CNPJ - Destinatário"],
            p["CNPJ da empresa que pagará o frete"]}
    assert len(tres) == 3


def test_cubagem_nao_vai_para_o_formulario():
    """DV não tem campo de m³. Cubagem é interna e não pode vazar no envio."""
    p = dv.preparar_payload(montar())
    campos = dv.campos_do_formulario(p)
    assert "cubagem_m3" in p["_meta"]
    assert not any("cubagem" in k.lower() or "m³" in k for k in campos)


def test_medidas_distintas_usam_o_maior_volume():
    r = montar(volumes=[
        Volume(qtd=1, comprimento_cm=Decimal(80), largura_cm=Decimal(40),
               altura_cm=Decimal(30), peso_kg=Decimal(8)),
        Volume(qtd=1, comprimento_cm=Decimal(100), largura_cm=Decimal(50),
               altura_cm=Decimal(40), peso_kg=Decimal(10)),
    ])
    p = dv.preparar_payload(r)
    assert (p["Comprimento"], p["Largura"], p["Altura"]) == ("100,0", "50,0", "40,0")
    assert p["Anexar Planilha"] == ["__PLANILHA_VOLUMES__"]


def test_ftl_inclui_tipo_de_veiculo():
    p = dv.preparar_payload(montar(
        servico=Servico.LOTACAO_FTL,
        veiculo_desejado="Truck (até 12.500 kg)",
    ))
    assert p["Qual o serviço que você procura?"] == "Lotação/Dedicado-FTL"
    assert p["Escolha o tipo de veículo"] == "Truck (até 12.500 kg)"


def test_ltl_nao_inclui_tipo_de_veiculo():
    assert "Escolha o tipo de veículo" not in dv.preparar_payload(montar())


# ---------------------------------------------------------------- validação
def test_carga_valida_nao_tem_bloqueio():
    assert dv.bloqueantes(dv.validar(montar())) == []


def test_peso_acima_de_34t_bloqueia():
    r = montar(volumes=[Volume(qtd=1, comprimento_cm=Decimal(100),
                               largura_cm=Decimal(100), altura_cm=Decimal(100),
                               peso_kg=Decimal(35_000))])
    erros = dv.bloqueantes(dv.validar(r))
    assert any(e.campo == "peso_total_kg" for e in erros)


def test_ftl_sem_veiculo_bloqueia():
    erros = dv.bloqueantes(dv.validar(montar(servico=Servico.LOTACAO_FTL)))
    assert any(e.campo == "veiculo_desejado" for e in erros)


def test_veiculo_pequeno_demais_bloqueia():
    r = montar(
        servico=Servico.LOTACAO_FTL,
        veiculo_desejado="Fiorino (até 500 kg)",
        volumes=[Volume(qtd=1, comprimento_cm=Decimal(100), largura_cm=Decimal(100),
                        altura_cm=Decimal(100), peso_kg=Decimal(2_000))],
    )
    erros = dv.bloqueantes(dv.validar(r))
    assert any("Fiorino" in e.mensagem for e in erros)


def test_quimico_sem_fispq_bloqueia():
    r = montar(mercadoria=Mercadoria(tipo_material="Soda cáustica", is_perigoso=True))
    assert any(e.campo == "mercadoria.fispq_path" for e in dv.bloqueantes(dv.validar(r)))


def test_medidas_distintas_e_aviso_nao_bloqueio():
    r = montar(volumes=[
        Volume(qtd=1, comprimento_cm=Decimal(100), largura_cm=Decimal(50),
               altura_cm=Decimal(40), peso_kg=Decimal(10)),
        Volume(qtd=1, comprimento_cm=Decimal(80), largura_cm=Decimal(40),
               altura_cm=Decimal(30), peso_kg=Decimal(8)),
    ])
    erros = dv.validar(r)
    assert dv.bloqueantes(erros) == []
    assert any(e.severidade is Severidade.AVISO for e in erros)


# ----------------------------------------------------------------- resposta
def test_envio_confirmado_vira_aguardando_retorno_sem_preco():
    """A frase de confirmação vem DENTRO do bloco do CF7 — e é o bloco que
    prova o envio, não a frase.

    Este teste aceitava a frase solta. Era o contrato antigo, e foi ele que
    deixou passar o falso positivo: a página da Della Volpe tem uma seção
    "Casos de sucesso", então qualquer HTML dela batia no critério."""
    res = dv.normalizar_resposta(
        '<div class="wpcf7-response-output" aria-hidden="true">'
        'Obrigado! Sua mensagem foi enviada com sucesso.</div>')
    assert res.status is StatusCotacao.AGUARDANDO_RETORNO
    assert res.valor_frete is None   # não inventar preço


def test_resposta_inesperada_vira_erro():
    assert dv.normalizar_resposta("<html>502 Bad Gateway</html>").status is StatusCotacao.ERRO


# ----------------------------------------------------------------- planilha
def test_planilha_reflete_os_volumes(tmp_path):
    r = montar(volumes=[
        Volume(qtd=3, comprimento_cm=Decimal(100), largura_cm=Decimal(50),
               altura_cm=Decimal(40), peso_kg=Decimal(10), descricao="Caixas"),
        Volume(qtd=1, comprimento_cm=Decimal(80), largura_cm=Decimal(40),
               altura_cm=Decimal(30), peso_kg=Decimal(8), descricao="Pallet"),
    ])
    caminho = gerar_planilha_volumes(r, tmp_path / "volumes.xlsx")
    ws = load_workbook(caminho).active
    assert ws.cell(row=5, column=2).value == "Caixas"
    assert ws.cell(row=5, column=8).value == 30.0        # 10 x 3
    assert ws.cell(row=7, column=3).value == 4           # total de volumes
    assert ws.cell(row=7, column=8).value == 38.0        # 30 + 8


# ------------------------------------------------------------------ anexos
def test_anexos_nao_vao_junto_com_campos_de_texto():
    """Regressão: caminho de arquivo passado ao fill() quebra o envio.
    Só acontecia com produto químico — o caso em que o anexo é obrigatório."""
    r = montar(mercadoria=Mercadoria(tipo_material="Soda cáustica",
                                     is_perigoso=True,
                                     fispq_path="/tmp/fispq.pdf"))
    campos = dv.campos_do_formulario(dv.preparar_payload(r))
    texto, arquivos = dv.separar_anexos(campos)

    assert "Anexar FISPQ / Licença" not in texto
    assert arquivos["Anexar FISPQ / Licença"] == "/tmp/fispq.pdf"
    assert all(not str(v).startswith("/") for v in texto.values())


def test_separar_anexos_isola_planilha():
    r = montar(volumes=[
        Volume(qtd=1, comprimento_cm=Decimal(100), largura_cm=Decimal(50),
               altura_cm=Decimal(40), peso_kg=Decimal(10)),
        Volume(qtd=1, comprimento_cm=Decimal(80), largura_cm=Decimal(40),
               altura_cm=Decimal(30), peso_kg=Decimal(8)),
    ])
    texto, arquivos = dv.separar_anexos(
        dv.campos_do_formulario(dv.preparar_payload(r)))
    assert "Anexar Planilha" not in texto
    assert "Anexar Planilha" in arquivos


def test_carga_sem_anexo_nao_cria_chave_de_arquivo():
    texto, arquivos = dv.separar_anexos(
        dv.campos_do_formulario(dv.preparar_payload(montar())))
    assert arquivos == {}
    assert len(texto) == 18
