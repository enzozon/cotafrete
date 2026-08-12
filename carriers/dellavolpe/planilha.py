"""Gera o .xlsx que a Della Volpe pede quando há volumes com medidas diferentes.

O site diz: "Se houver mais volumes com medidas diferentes, anexar planilha em
Excel, informando pesos e medidas de todos os volumes". Em vez de exigir isso do
usuário, o adapter monta a planilha a partir da lista de volumes.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.models import CotacaoRequest

CABECALHO = [
    "Item", "Descrição", "Qtd",
    "Comprimento (cm)", "Largura (cm)", "Altura (cm)",
    "Peso unit. (kg)", "Peso total (kg)", "Cubagem (m³)",
]


def gerar_planilha_volumes(req: CotacaoRequest, destino: str | Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Volumes"

    ws["A1"] = "Relação de Volumes — Cotação de Frete"
    ws["A1"].font = Font(size=13, bold=True)
    ws.merge_cells("A1:I1")

    ws["A2"] = f"{req.origem.cidade}/{req.origem.uf}  →  {req.destino.cidade}/{req.destino.uf}"
    ws.merge_cells("A2:I2")

    fill = PatternFill("solid", fgColor="F8981D")
    for col, titulo in enumerate(CABECALHO, start=1):
        c = ws.cell(row=4, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    linha = 5
    for i, v in enumerate(req.volumes, start=1):
        ws.cell(row=linha, column=1, value=i)
        ws.cell(row=linha, column=2, value=v.descricao or f"Volume {i}")
        ws.cell(row=linha, column=3, value=v.qtd)
        ws.cell(row=linha, column=4, value=float(v.comprimento_cm))
        ws.cell(row=linha, column=5, value=float(v.largura_cm))
        ws.cell(row=linha, column=6, value=float(v.altura_cm))
        ws.cell(row=linha, column=7, value=float(v.peso_kg))
        ws.cell(row=linha, column=8, value=float(v.peso_total_kg))
        ws.cell(row=linha, column=9, value=round(float(v.cubagem_m3), 4))
        linha += 1

    ws.cell(row=linha, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=linha, column=3, value=req.quantidade_volumes).font = Font(bold=True)
    ws.cell(row=linha, column=8, value=float(req.peso_total_kg)).font = Font(bold=True)
    ws.cell(row=linha, column=9, value=round(float(req.cubagem_m3), 4)).font = Font(bold=True)

    for col, largura in zip("ABCDEFGHI", (6, 28, 8, 16, 14, 12, 14, 14, 13)):
        ws.column_dimensions[col].width = largura

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino
